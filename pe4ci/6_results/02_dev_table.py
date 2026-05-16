# %%
import math
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

from crisp.library import start

# %%
# ------------------ SETUP ------------------
CONCEPTS = ["gratitude", "ncb", "mm"]
PLATFORMS = ["openai", "llama3.3"]

TECHNIQUES = ["Baseline", "APE", "Persona", "Chain-of-Thought", "Explanations"]

TECHNIQUE_KEY_MAP = {
    ("Baseline", "Zero-Shot"): "baseline_zero",
    ("Baseline", "Few-Shot"): "baseline_few",
    ("APE", "Zero-Shot"): "ape_zero",
    ("APE", "Few-Shot"): "ape_few",
    ("Persona", "Zero-Shot"): "persona_zero",
    ("Persona", "Few-Shot"): "persona_few",
    ("Chain-of-Thought", "Zero-Shot"): "cot_zero",
    ("Chain-of-Thought", "Few-Shot"): "cot_few",
    ("Explanations", "Zero-Shot"): None,
    ("Explanations", "Few-Shot"): "explanation_few",
}

COLUMN_MAP = {
    "bottom": {"Zero-Shot": 2, "Few-Shot": 3},
    "top": {"Zero-Shot": 4, "Few-Shot": 5},
}


def get_best_row_for_category(technique_df, category):
    technique_df = technique_df.copy()

    if "category" not in technique_df.columns or technique_df["category"].isna().all():
        prompt_ids_contain_category = any(
            "top" in str(prompt_id).lower() or "bottom" in str(prompt_id).lower()
            for prompt_id in technique_df["prompt_id"]
        )

        if prompt_ids_contain_category:
            technique_df["category"] = technique_df["prompt_id"].apply(
                lambda prompt_id: "top" if "top" in str(prompt_id).lower() else "bottom"
            )
        else:
            technique_df["category"] = technique_df["F1"].apply(
                lambda f1: "top" if f1 == technique_df["F1"].max() else "bottom"
            )

    category_df = technique_df[technique_df["category"] == category].sort_values(
        "F1", ascending=False
    )

    if category_df.empty:
        return None

    return category_df.iloc[0]


# %%
# ------------------ BUILD DEV TABLES ------------------
for concept in CONCEPTS:
    for platform in PLATFORMS:
        dev_results_path = start.RESULTS_DIR + f"{platform}_{concept}_dev_results.xlsx"
        output_path = start.RESULTS_DIR + f"dev_table_{platform}_{concept}.xlsx"

        if not os.path.exists(dev_results_path):
            print(f"Skipping {platform} - {concept}: {dev_results_path} not found")
            continue

        print(f"Processing {platform} - {concept}...")

        all_results_df = pd.read_excel(dev_results_path, sheet_name="results")

        if os.path.exists(output_path):
            wb = load_workbook(output_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "summary"

            ws.cell(row=1, column=2, value="Bottom Baseline")
            ws.cell(row=2, column=2, value="Zero-Shot")
            ws.cell(row=2, column=3, value="Few-Shot")
            ws.cell(row=1, column=4, value="Top Baseline")
            ws.cell(row=2, column=4, value="Zero-Shot")
            ws.cell(row=2, column=5, value="Few-Shot")

            label_row = 3
            for technique in TECHNIQUES:
                ws.cell(row=label_row, column=1, value=technique)
                label_row += 2

        baseline_df = all_results_df[all_results_df["technique"] == "baseline_zero"]

        baseline_reference = {}
        for category in ["top", "bottom"]:
            best_row = get_best_row_for_category(baseline_df, category)

            if best_row is None:
                baseline_reference[category] = None
            else:
                baseline_reference[category] = {
                    "F1": best_row["F1"],
                    "F1 SE": best_row["F1 SE"],
                }

        for strategy in ["Zero-Shot", "Few-Shot"]:
            for category in ["top", "bottom"]:
                excel_column = COLUMN_MAP[category][strategy]
                excel_row = 3

                for technique in TECHNIQUES:
                    technique_key = TECHNIQUE_KEY_MAP.get((technique, strategy))

                    if technique_key is None:
                        excel_row += 2
                        continue

                    technique_df = all_results_df[all_results_df["technique"] == technique_key]

                    if technique_df.empty:
                        print(f"No data found for {platform} - {concept} - {technique} {strategy}")
                        excel_row += 2
                        continue

                    best_row = get_best_row_for_category(technique_df, category)

                    if best_row is None:
                        excel_row += 2
                        continue

                    f1 = best_row["F1"]
                    f1_se = best_row["F1 SE"]

                    significance_stars = ""
                    if baseline_reference[category] is not None:
                        base_f1 = baseline_reference[category]["F1"]
                        base_se = baseline_reference[category]["F1 SE"]
                        pooled_se = math.sqrt(f1_se**2 + base_se**2)
                        z_score = abs(f1 - base_f1) / pooled_se if pooled_se > 0 else 0

                        if z_score > 2.58:
                            significance_stars = "***"
                        elif z_score > 1.96:
                            significance_stars = "**"
                        elif z_score > 1.64:
                            significance_stars = "*"

                    ws.cell(row=excel_row, column=excel_column, value=f"{f1:.2f}{significance_stars}")
                    ws.cell(row=excel_row + 1, column=excel_column, value=f"({f1_se:.2f})")

                    excel_row += 2

        wb.save(output_path)
        print(f"Saved: {output_path}")
