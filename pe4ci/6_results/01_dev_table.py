import math
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import sys

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.append(PROJECT_ROOT)
from pe4ci.library import start

# ------------------ SETUP ------------------
# CONCEPT = start.CONCEPT
# PLATFORM = start.PLATFORM

CONCEPT = "ncb"
PLATFORM = "openai"

RESULTS_FILE = start.RESULTS_DIR + f"dev_table_{PLATFORM}_{CONCEPT}.xlsx"

TECHNIQUES = ["Baseline", "APE", "Persona", "Chain-of-Thought", "Explanations", "Cloze"]
RESULT_PATHS = {
    "Baseline": {
        "Zero-Shot": f"{PLATFORM}_{CONCEPT}_baseline_zero_results_dev.xlsx",
        "Few-Shot": f"{PLATFORM}_{CONCEPT}_baseline_few_results_dev.xlsx",
        # "Fine-Tuning": ...
    },
    "APE": {
        "Zero-Shot": f"{PLATFORM}_{CONCEPT}_ape_zero_results_dev.xlsx",
        "Few-Shot": f"{PLATFORM}_{CONCEPT}_ape_few_results_dev.xlsx",
    },
    "Persona": {
        "Zero-Shot": f"{PLATFORM}_{CONCEPT}_persona_zero_results_dev.xlsx",
        "Few-Shot": f"{PLATFORM}_{CONCEPT}_persona_few_results_dev.xlsx",
    },
    "Chain-of-Thought": {
        "Zero-Shot": f"{PLATFORM}_{CONCEPT}_cot_zero_results_dev.xlsx",
        "Few-Shot": f"{PLATFORM}_{CONCEPT}_cot_few_results_dev.xlsx",
    },
    "Explanations": {
        "Zero-Shot": f"{PLATFORM}_{CONCEPT}_explanation_zero_results_dev.xlsx",
        "Few-Shot": f"{PLATFORM}_{CONCEPT}_explanation_few_results_dev.xlsx",
    },
}
# ------------------ SETUP WORKBOOK ------------------
if os.path.exists(RESULTS_FILE):
    wb = load_workbook(RESULTS_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.cell(row=1, column=2, value="Bottom Baseline")
    ws.cell(row=2, column=2, value="Zero-Shot")
    ws.cell(row=2, column=3, value="Few-Shot")
    ws.cell(row=1, column=4, value="Top Baseline")
    ws.cell(row=2, column=4, value="Zero-Shot")
    ws.cell(row=2, column=5, value="Few-Shot")

    row = 3
    for technique in TECHNIQUES:
        ws.cell(row=row, column=1, value=technique)
        row += 2
wb.save(RESULTS_FILE)

# ------------------ COLLECT BASELINE VALUES FOR SIGNIFICANCE ------------------
baseline_values = {}

for category in ["top", "bottom"]:
    file_path = os.path.join(
        start.MAIN_DIR, "results", RESULT_PATHS["Baseline"]["Zero-Shot"]
    )
    df = pd.read_excel(file_path, sheet_name="results")

    if "category" not in df.columns:
        top_or_bottom_in_prompt_id = any(
            ("top" in str(pid).lower() or "bottom" in str(pid).lower())
            for pid in df["prompt_id"]
        )
        if top_or_bottom_in_prompt_id:
            df["category"] = df["prompt_id"].apply(
                lambda pid: "top" if "top" in str(pid).lower() else "bottom"
            )
        else:
            df["category"] = df["F1"].apply(
                lambda f1: "top" if f1 == df["F1"].max() else "bottom"
            )

    cat_df = df[df["category"] == category].sort_values("F1", ascending=False).head(1)
    baseline_values[category] = {
        "F1": cat_df["F1"].values[0],
        "SE": cat_df["F1 SE"].values[0],
    }

# ------------------ WRITE ONLY TOP ZERO-SHOT ------------------
start_row = 3
column_maps = {
    "top": {
        "Zero-Shot": 4,
        "Few-Shot": 5,
    },
    "bottom": {
        "Zero-Shot": 2,
        "Few-Shot": 3,
    },
}

for strategy in ["Zero-Shot", "Few-Shot"]:
    for category in ["bottom", "top"]:
        col = column_maps[category][strategy]
        row = start_row
        for technique in TECHNIQUES:
            if technique not in RESULT_PATHS:
                row += 2
                continue
            file_path = os.path.join(
                start.MAIN_DIR, "results", RESULT_PATHS[technique][strategy]
            )

            try:
                df = pd.read_excel(file_path, sheet_name="results")
            except Exception as e:
                print(f"Error loading {technique}: {e}")
                row += 2
                continue

            if "category" not in df.columns:
                top_or_bottom_in_prompt_id = any(
                    ("top" in str(pid).lower() or "bottom" in str(pid).lower())
                    for pid in df["prompt_id"]
                )
                if top_or_bottom_in_prompt_id:
                    df["category"] = df["prompt_id"].apply(
                        lambda pid: "top" if "top" in str(pid).lower() else "bottom"
                    )
                else:
                    df["category"] = df["F1"].apply(
                        lambda f1: "top" if f1 == df["F1"].max() else "bottom"
                    )

            cat_df = (
                df[df["category"] == category]
                .sort_values("F1", ascending=False)
                .head(1)
            )
            if cat_df.empty:
                row += 2
                continue

            f1 = cat_df["F1"].values[0]
            se = cat_df["F1 SE"].values[0]

            # Significance testing vs. baseline
            base_f1 = baseline_values[category]["F1"]
            base_se = baseline_values[category]["SE"]
            pooled_se = math.sqrt(se**2 + base_se**2)
            z = abs(f1 - base_f1) / pooled_se if pooled_se > 0 else 0

            if z > 2.58:
                stars = "***"
            elif z > 1.96:
                stars = "**"
            elif z > 1.64:
                stars = "*"
            else:
                stars = ""

            ws.cell(row=row, column=col, value=f"{f1:.2f}{stars}")
            ws.cell(row=row + 1, column=col, value=f"({se:.2f})")
            row += 2

# ------------------ SAVE ------------------
wb.save(RESULTS_FILE)
