# classify.py
import os
import numpy as np
import pandas as pd
from tqdm import tqdm
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openpyxl import Workbook, load_workbook
import re

from crisp.library import start, metric_standard_errors
from crisp.library.start import PLATFORM

if PLATFORM == "openai":
    from crisp.library import secrets


# ------------------ CLIENT SETUP ------------------
if "llama" in start.PLATFORM:
    from langchain_ollama import OllamaLLM

if start.PLATFORM == "openai":
    from openai import OpenAI

    OPENAI_API_KEY = secrets.OPENAI_API_KEY
    client = OpenAI(api_key=OPENAI_API_KEY)

ollama_server_url = "http://localhost:11434"


# ------------------ RESPONSE FUNCTIONS ------------------
def get_model_response(model_provider, message_text, temperature=0.0001):
    if model_provider == "openai":
        response = client.responses.create(
            model=start.MODEL,
            input=message_text,
            reasoning={"effort": "none"},
        )
        cleaned_response = response.output_text.strip() if response.output_text else ""
        return cleaned_response, "fingerprint n/a"

    elif model_provider.startswith("llama"):
        llm = OllamaLLM(
            model=start.MODEL,
            base_url=ollama_server_url,
            temperature=temperature,

            seed=start.SEED,
        )
        response = llm.invoke(message_text)
        return response, "fingerprint n/a"

    else:
        raise ValueError(f"Unsupported model provider: {model_provider}")


def parse_binary_classification(response_text: str):
    """
    Parse a yes/no classification from the model response.

    Priority:
    1. Triple-backticked yes/no
    2. Any backticked yes/no
    3. Whole-word yes/no fallback
    """
    if pd.isna(response_text) or not str(response_text).strip():
        return np.nan

    text = str(response_text)

    triple_matches = re.findall(r"`{3}\s*(yes|no)\s*`{3}", text, flags=re.IGNORECASE)
    if triple_matches:
        return 1 if triple_matches[-1].lower() == "yes" else 0

    backtick_matches = re.findall(r"`+\s*(yes|no)\s*`+", text, flags=re.IGNORECASE)
    if backtick_matches:
        return 1 if backtick_matches[-1].lower() == "yes" else 0

    word_matches = re.findall(r"\b(yes|no)\b", text, flags=re.IGNORECASE)
    if word_matches:
        return 1 if word_matches[-1].lower() == "yes" else 0

    return np.nan


# ------------------ PROMPT EVALUATION ------------------
def get_classifications_from_prompt(
    prompt_text,
    prompt_id,
    df,
    platform,
    temperature=0.0001,
):
    """
    prompt_text must contain the literal substring {TEXT}.
    That placeholder will be replaced with each row's text.
    """
    rows = []

    if "{TEXT}" not in prompt_text:
        raise ValueError("prompt_text must contain the placeholder {TEXT}")

    for unique_text_id, text in tqdm(
        zip(df.unique_text_id, df.text),
        total=len(df),
        desc=f"Prompt {prompt_id}",
        position=0,
        leave=False,
    ):
        full_message = prompt_text.replace("{TEXT}", str(text))

        cleaned_response, system_fingerprint = get_model_response(
            model_provider=platform,
            message_text=full_message,
            temperature=temperature,
        )

        classification = parse_binary_classification(cleaned_response)

        rows.append(
            {
                "unique_text_id": unique_text_id,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt_text,
                "model": start.MODEL,
                "fingerprint": system_fingerprint,
                "prompt_id": prompt_id,
            }
        )

    return rows


# ------------------ PROMPT VARIATION ------------------
def generate_prompt_variants(
    model_provider, base_prompt, metaprompt1, metaprompt2, num_variants
):
    variants = []

    for _ in range(num_variants):
        meta_instructions = metaprompt1 + base_prompt + metaprompt2

        if model_provider == "openai":
            response = client.responses.create(
                model=start.MODEL,
                input=meta_instructions,
                reasoning={"effort": "none"},
            )
            new_prompt = response.output_text.strip() if response.output_text else ""

        elif model_provider.startswith("llama"):
            llm = OllamaLLM(
                model=start.MODEL,
                base_url=ollama_server_url,
                temperature=1,
            )
            new_prompt = llm.invoke(meta_instructions)

        else:
            raise ValueError(f"Unsupported model provider: {model_provider}")

        variants.append(new_prompt)

    return variants


# ------------------ EXCEL EXPORT ------------------
def export_results_to_excel(
    df,
    output_path,
    group_col="prompt_id",
    prompt_col="prompt",
    y_true_col="human_code",
    y_pred_col="classification",
    sheet_name="results",
    include_se=True,
    n_bootstraps=1000,
    random_state=12,
):
    if not os.path.exists(output_path):
        workbook = Workbook()
        workbook.save(output_path)

    workbook = load_workbook(output_path)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)

    if isinstance(group_col, str):
        group_col = [group_col]

    headers = group_col + ["Accuracy", "Precision", "Recall", "F1"]
    if include_se:
        headers += ["Accuracy SE", "Precision SE", "Recall SE", "F1 SE"]
    headers.append(prompt_col)

    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    grouped = df.dropna(subset=[y_true_col, y_pred_col]).groupby(group_col)

    row = 2
    for group_vals, group_df in grouped:
        y_true = group_df[y_true_col]
        y_pred = group_df[y_pred_col]
        prompt_text = group_df[prompt_col].iloc[0]

        accuracy = accuracy_score(y_true, y_pred)
        precision = precision_score(y_true, y_pred)
        recall = recall_score(y_true, y_pred)
        f1 = f1_score(y_true, y_pred)
        print(f"Accuracy: {accuracy}")
        print(f"Precision: {precision}")
        print(f"Recall: {recall}")
        print(f"F1: {f1}")
        result_row = list(group_vals) if isinstance(group_vals, tuple) else [group_vals]
        result_row += [accuracy, precision, recall, f1]

        if include_se:
            participant_ids = group_df["unique_text_id"].str.rsplit("_", n=1).str[0]
            _, accuracy_se = metric_standard_errors.bootstrap_accuracy(
                y_true, y_pred, participant_ids, n_bootstraps, random_state
            )
            _, precision_se = metric_standard_errors.bootstrap_precision(
                y_true, y_pred, participant_ids, n_bootstraps, random_state
            )
            _, recall_se = metric_standard_errors.bootstrap_recall(
                y_true, y_pred, participant_ids, n_bootstraps, random_state
            )
            _, f1_se = metric_standard_errors.bootstrap_f1(
                y_true, y_pred, participant_ids, n_bootstraps, random_state
            )
            result_row += [accuracy_se, precision_se, recall_se, f1_se]

        result_row.append(prompt_text)

        for col_index, val in enumerate(result_row, 1):
            worksheet.cell(
                row=row,
                column=col_index,
                value=round(val, 3) if isinstance(val, float) else val,
            )
        row += 1

    workbook.save(output_path)

