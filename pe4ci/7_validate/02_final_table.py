# %%
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

from crisp.library import start

# %%
# ------------------ SETUP ------------------
CONCEPTS = ["gratitude", "mm", "ncb"]
PLATFORM = "openai"

RESULTS_FILE = start.RESULTS_DIR + "final_test_results.xlsx"

# %%
# ------------------ CREATE/LOAD FINAL WORKBOOK ------------------
if os.path.exists(RESULTS_FILE):
    wb = load_workbook(RESULTS_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

for row in ws.iter_rows():
    for cell in row:
        cell.value = None

ws.cell(row=1, column=1, value="Concept")
ws.cell(row=1, column=2, value="Platform")
ws.cell(row=1, column=3, value="Technique")
ws.cell(row=1, column=4, value="Category")
ws.cell(row=1, column=5, value="Accuracy")
ws.cell(row=1, column=6, value="Specificity")
ws.cell(row=1, column=7, value="Recall")
ws.cell(row=1, column=8, value="Precision")
ws.cell(row=1, column=9, value="F1")
ws.cell(row=1, column=10, value="NN Rate")
ws.cell(row=1, column=11, value="Prompt")

# %%
# ------------------ FILL TABLE ------------------
excel_row = 2

for concept in CONCEPTS:
    df = pd.read_excel(start.RESULTS_DIR + f"{PLATFORM}_{concept}_test_results.xlsx", sheet_name="results")
    best_row = df.iloc[0]

    ws.cell(row=excel_row, column=1, value=concept)
    ws.cell(row=excel_row, column=2, value=PLATFORM)
    ws.cell(row=excel_row, column=3, value=best_row["technique"])
    ws.cell(row=excel_row, column=4, value=best_row["category"])

    for col, metric in zip(range(5, 11), ["Accuracy", "Specificity", "Recall", "Precision", "F1", "NN Rate"]):
        ws.cell(row=excel_row, column=col, value=round(best_row[metric], 2))

    ws.cell(row=excel_row, column=11, value=best_row["prompt"])

    excel_row += 1

    ws.cell(row=excel_row, column=4, value="95% CI")

    for col, metric in zip(range(5, 11), ["Accuracy", "Specificity", "Recall", "Precision", "F1", "NN Rate"]):
        ci_lower = round(best_row[f"{metric} CI Lower"], 2)
        ci_upper = round(best_row[f"{metric} CI Upper"], 2)
        ws.cell(row=excel_row, column=col, value=f"[{ci_lower:.2f}, {ci_upper:.2f}]")

    excel_row += 1

# %%
# ------------------ SAVE ------------------
wb.save(RESULTS_FILE)
print(f"Saved: {RESULTS_FILE}")
